"""
This module takes care of starting the API Server, Loading the DB and Adding the endpoints
"""
import os
import shutil
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
from werkzeug.utils import secure_filename

from flask import Flask, request, jsonify, url_for, json, send_from_directory
from flask_migrate import Migrate
from flask_swagger import swagger
from flask_cors import CORS
from flask_jwt_extended import create_access_token, JWTManager
from utils import APIException, generate_sitemap
from admin import setup_admin
from models import db, User, CommonData, Professor, Cathedra, CathedraAssign, Student, Course, Inscription, Grade, Evaluation, Career

#UPLOAD_FOLDER = '/files'

app = Flask(__name__, static_folder="static")
app.url_map.strict_slashes = False
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DB_CONNECTION_STRING')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config["JWT_SECRET_KEY"] = "01cdeef14f0a17d28d723f35a2ba3670"
app.config['UPLOAD_FOLDER'] = '/files'

MIGRATE = Migrate(app, db)
db.init_app(app)
CORS(app)
setup_admin(app)
jwt = JWTManager(app)

# Handle/serialize errors like a JSON object
@app.errorhandler(APIException)
def handle_invalid_usage(error):
    return jsonify(error.to_dict()), error.status_code

# generate sitemap with all your endpoints
@app.route('/')
def sitemap():
    return generate_sitemap(app)

@app.route("/sign-up", methods=["POST"])
def sign_up():
    data = request.json

    if 'professor_id' in data.keys():
        user = User.create(email=data.get('email'), password=data.get('password'), role=data.get('role'), professor_id=data.get('professor_id'))
    else :
        user = User.create(email=data.get('email'), password=data.get('password'), role=data.get('role'))

    if not isinstance(user, User):
        return jsonify({"msg": "Ocurrio un problema interno"}), 500

    return jsonify(user.serialize()), 201

@app.route("/log-in", methods=["POST"])
def log_in():
    data = request.json
    user = User.query.filter_by(email=data['email']).one_or_none()

    if user is None: 
        return jsonify({"msg": "No existe el usuario"}), 404
    if not user.check_password(data.get('password')):
        return jsonify({"msg": "Credenciales incorrectas"}), 400

    token = create_access_token(identity=user.id)

    return jsonify({
        "user": user.serialize(),
        "token": token
    }), 200

# endpoint for getting the possibles career
@app.route("/careers", methods=["GET"])
def get_all_careers():
    '''
        Get all careers
    '''
    careers = []
    for career in Career:
        careers.append(career.name)

    return jsonify(careers), 200

@app.route("/careers/info", methods=["GET"])
def get_careers_info():
    '''
        Get the info of all the careers
    '''
    careers_info = []
    for career in Career:
        career_info = {}
        # career_info = {
        #     "cathedras": [],
        #     "professors": [],
        #     "students": []
        # }
        # query all the information for the cathedra
        cathedras_in_career = Cathedra.query.filter_by(career=career.name).all()
        professors_in_career = Professor.query.filter_by(career=career.name).all()
        students_in_career = Student.query.filter_by(career=career.name).all()
        # assign it to the dictionary
        career_info["name"] = career.name
        career_info["cathedras"] = [cathedra.serialize() for cathedra in cathedras_in_career]
        career_info["professors"] = [professor.serialize() for professor in professors_in_career]
        career_info["students"] = [student.serialize() for student in students_in_career]
        careers_info.append(career_info)
    
    return jsonify(careers_info), 200

@app.route("/get-careers-file", methods=["POST"])
def get_careers_info_file():
    '''
        Get all careers info
    '''
    careers_info = []
    for career in Career:
        career_info = {}
        # career_info = {
        #     "cathedras": [],
        #     "professors": [],
        #     "students": []
        # }
        # query all the information for the cathedra
        cathedras_in_career = Cathedra.query.filter_by(career=career.name).all()
        professors_in_career = Professor.query.filter_by(career=career.name).all()
        students_in_career = Student.query.filter_by(career=career.name).all()
        # assign it to the dictionary
        career_info["name"] = career.name
        career_info["cathedras"] = cathedras_in_career
        career_info["professors"] = professors_in_career
        career_info["students"] = students_in_career
        careers_info.append(career_info)

    file_name = "InfoCarreras"
    wb = Workbook()
    ws = wb.active
    ws.title = file_name
    
    for title, info in {"A1": "Carrera", "B1": "Cantidad de materias","C1": "Cantidad de profesores", "D1": "Cantidad de estudiantes"}.items():
        ws[title] = info
    
    i = 2 
    for career_info in careers_info:
        aux = ['A'+str(i), 'B'+str(i), 'C'+str(i), 'D'+str(i)]
        ws[aux[0]] = career_info["name"]
        ws[aux[1]] = len(career_info["cathedras"])
        ws[aux[2]] = len(career_info["professors"])
        ws[aux[3]] = len(career_info["students"])
        i += 1

    wb.save(file_name+'.xlsx')
    shutil.move(file_name+'.xlsx', "src/static/reports/")

    return jsonify({"msg": "Archivo generado", "file_name": file_name}), 200

@app.route("/static-file/<file_name>", methods=["GET"])
def create_static_image(file_name):

    secured_filename = secure_filename(file_name+'.xlsx')
    image_path = os.path.join("reports", secured_filename)
    print(secured_filename, image_path)
    print(app.static_folder)
    print(os.path.join(app.static_folder, image_path))
    print(os.path.exists(os.path.join(app.static_folder, image_path)))
    
    if os.path.exists(os.path.join(app.static_folder, image_path)):
        return send_from_directory(app.static_folder, image_path)
    else:
        return jsonify({"msg": "Error archivo no encontrado"}), 404

# endpoints for cathedra
@app.route("/cathedra", methods=["GET"])
def get_all_cathedras():
    '''
        Get all cathedras
    '''
    cathedras = [cathedra.serialize() for cathedra in Cathedra.query.all()]
    return jsonify(cathedras), 200

@app.route("/cathedra", methods=["POST"])
def create_cathedra():
    '''
        Creates a new cathedra with just "name, code, credits, career"
    '''
    data = json.loads(request.data)
    new_cathedra = Cathedra(
        name=data["name"],
        code=data["code"],
        credits=data["credits"],
        career=data["career"]
    )

    db.session.add(new_cathedra)
    try:
        db.session.commit()
    except:
        db.session.rollback()
        return jsonify({"msg": "Hubo un error creando la materia"}), 500

    return jsonify(new_cathedra.serialize()), 200

@app.route("/cathedras/<int:career>", methods=["GET"])
def get_cathedras_of_career(career):
    '''
        Return all the cathedras of a career
    '''

    cathedras = Cathedra.query.filter_by(career=Career(career).name)

    return jsonify([cathedra.serialize() for cathedra in cathedras]), 200

@app.route("/upload-cathedras", methods=["POST"])
def upload_cathedras_file():
    '''
        Upload the data of cathedras from file
    '''
    # wb = workbook 
    try:   
        myFile = request.files["myFile"]
        wb = load_workbook(myFile)    
    except: 
        return jsonify({"msg": "Hubo un problema abriendo el archivo"}), 500

    for sheet_name in wb.sheetnames:
        # ws = worksheet
        ws = wb[sheet_name]
        # the ws is a dictionary but the rows are tuples
        for row in ws.iter_rows(min_row=2):
            new_cathedra = Cathedra(
                name=row[0].value,
                code=row[1].value,
                credits=row[2].value,
                career=row[3].value
            )
            db.session.add(new_cathedra)
    
    try:
        db.session.commit()
    except:
        db.session.rollback()
        return jsonify({"msg": "Hubo un problema creando la catedra"}), 500

    return jsonify({"msg": "Se anadireron las catedras del archivo"}), 200

# endpoints for professor
@app.route("/professor", methods=["POST"])
def create_professor():
    '''
        Creates a professor with its user
    '''
    data = request.json
    # create the new professor
    new_professor = Professor(
        full_name=data["fullName"],
        ci=data["ci"],
        phone_number=data["phoneNumber"],
        age=data["age"],
        nationality=data["nationality"],
        residence=data["residence"],
        career=data["career"]
    )
    db.session.add(new_professor)
    try:
        db.session.commit()
    except:
        db.session.rollback()
        return jsonify({"msg": "Hubo un error creando al profesor"}), 500

    # loop trough the cathedras codes
    for cathedra_code in data["cathedras"]:
        cathedra = Cathedra.query.filter_by(code=cathedra_code).all()[0]
        # create the relation
        new_relation = CathedraAssign(professor_id=new_professor.id, cathedra_id=cathedra.id)
        db.session.add(new_relation)
        # if the user is a coordinator create the relation with the cathedra
        if data["role"] == 2:
            cathedra.coordinator_id = new_professor.id

    try:
        db.session.commit()
    except:
        db.session.rollback()
        return jsonify({"msg": "Hubo un error creando las relaciones"}), 500
    
    # create the user with the professor id
    new_user = User(
        email=data['email'],
        password=data['ci'], 
        role=data['role'], 
        professor_id=new_professor.id
    )
    db.session.add(new_user)

    try:
        db.session.commit()
    except:
        db.session.rollback()
        return jsonify("Hubo un problema creando el usuario"), 500 

    return jsonify({
            "user": new_user.serialize()
        }), 200

@app.route("/professors/info", methods=["GET"])
def get_professors_info():
    '''
        Get the info of all the careers
    '''
    professors_info = Professor.query.all()
    
    return jsonify([professor.serialize() for professor in professors_info]), 200

@app.route("/upload-professors", methods=["POST"])
def upload_professors_file():
    '''
        Upload the data of professors from file
    '''
    # wb = workbook 
    try:   
        myFile = request.files["myFile"]
        wb = load_workbook(myFile)    
    except: 
        return jsonify({"msg": "Hubo un problema abriendo el archivo"}), 500

    for sheet_name in wb.sheetnames:
        # ws = worksheet
        ws = wb[sheet_name]
        # the ws is a dictionary but the rows are tuples
        for row in ws.iter_rows(min_row=2):
            new_professor = Professor(
                full_name=row[0].value,
                ci=row[1].value,
                phone_number=row[2].value,
                age=row[3].value,
                nationality=row[4].value,
                residence=row[5].value,
                career=row[6].value
            )
            db.session.add(new_professor)
            # save the professor to get the id
            try:
                db.session.commit()
            except:
                db.session.rollback()
                return jsonify({"msg": "Hubo un problema creando el profesor"}), 500

            # creating the list of codes
            cathedras_codes = row[7].value[1:-1].split(',')
            # creating the relations
            for cathedra_code in cathedras_codes:
                cathedra = Cathedra.query.filter_by(code=cathedra_code).all()[0]
                # creates the relation
                new_relation = CathedraAssign(professor_id=new_professor.id, cathedra_id=cathedra.id)
                db.session.add(new_relation)
                # if the user is a coordinator create the relation with the cathedra
                if row[9].value == 2:
                    cathedra.coordinator_id = new_professor.id

            try:
                db.session.commit()
            except: 
                db.session.callback()
                return jsonify({"msg": "Hubo un error creando las relaciones"}), 500

            # create the user with the professor id
            new_user = User(
                email=row[8].value,
                password=row[1].value, 
                role=row[9].value, 
                professor_id=new_professor.id
            )
            db.session.add(new_user)

            try:
                db.session.commit()
            except:
                db.session.rollback()
                return jsonify("Hubo un problema creando el usuario"), 500 

    return jsonify({"msg": "Se anadireron los profesores del archivo"}), 200

# endpoints for student
@app.route("/student", methods=["POST"])
def create_student():
    '''
        Create a student
    '''
    data = request.json
    new_student = Student(
        full_name=data["fullName"],
        ci=data["ci"],
        phone_number=data["phoneNumber"],
        age=data["age"],
        nationality=data["nationality"],
        residence=data["residence"],
        career=data["career"]
    )
    db.session.add(new_student)

    try:
        db.session.commit()
    except:
        db.session.rollback()
        return jsonify({"msg": "Hubo un error creando el estudiante"}), 500
    
    return jsonify(new_student.serialize()), 200

@app.route("/students/info", methods=["GET"])
def get_students_info():
    '''
        Get the info of all the careers
    '''
    students_info = Student.query.all()
    
    return jsonify([student.serialize() for student in students_info]), 200

@app.route("/upload-students", methods=["POST"])
def upload_students_file():
    '''
        Upload the data of students from file
    '''
    # wb = workbook 
    try: 
        myFile = request.files["myFile"]
        wb = load_workbook(myFile)
    except: 
        return jsonify({"msg": "Hubo un problema abriendo el archivo"}), 500
    
    for sheet_name in wb.sheetnames:
        # ws = worksheet
        ws = wb[sheet_name]
        # the ws is a dictionary but the rows are tuples
        for row in ws.iter_rows(min_row=2):
            new_student = Student(
                full_name=row[0].value,
                ci=row[1].value,
                phone_number=row[2].value,
                age=row[3].value,
                nationality=row[4].value,
                residence=row[5].value,
                career=row[6].value
            )
            db.session.add(new_student)

            try:
                db.session.commit()
            except:
                db.session.rollback()
                return jsonify({"msg": "Hubo un problema creando al estudiante"}), 500

            # creating the list of codes
            courses_codes = row[7].value[1:-1].split(',')
            # creating the inscriptions
            for course_code in courses_codes:
                course = Course.query.filter_by(code=course_code).all()[0]

                new_inscription = Inscription(student_id=new_student.id, course_id=course.id)
                db.session.add(new_inscription)

                try:
                    db.session.commit()
                except:
                    db.session.rollback()
                    return jsonify({"msg": "Hubo un problema creando la inscripcion"}), 500

    return jsonify({"msg": "Se anadireron los estudiantes del archivo"}), 200

# endpoints for inscriptions
@app.route("/inscription", methods=["POST"])
def create_inscription():
    '''
        Creates a inscription
    '''
    data = request.json
    new_inscription = Inscription(student_id=data["student_id"], course_id=data["course_id"])
    db.session.add(new_inscription)

    try:
        db.session.commit()
    except:
        db.session.rollback()
        return jsonify({"msg": "Hubo un error creando la inscripcion"}), 500
    
    return jsonify(new_inscription.serialize()), 200

# endpoints for courses
@app.route("/courses/<int:career>", methods=["GET"])
def get_courses_career(career):
    '''
        Get all the active courses from a career
    '''
    cathedras = Cathedra.query.filter_by(career=Career(career).name).all()
    courses = []
    for cathedra in cathedras:
        if cathedra.courses:
            for course in cathedra.courses:
                if course.is_active:
                    courses.append(course)

    return jsonify([course.serialize() for course in courses]), 200

@app.route("/courses/byCode/<code>", methods=["GET"])
def get_course_by_code(code):
    '''
        Get a course by its code
    '''
    course = Course.query.filter_by(code=str(code)).all()[0]

    return jsonify(course.serialize()), 200


@app.route("/upload-courses", methods=["POST"])
def upload_courses_file():
    '''
        Upload the data of courses from file
    '''
    # wb = workbook 
    try: 
        myFile = request.files["myFile"]
        wb = load_workbook(myFile)
    except: 
        return jsonify({"msg": "Hubo un problema abriendo el archivo"}), 500
    
    for sheet_name in wb.sheetnames:
        # ws = worksheet
        ws = wb[sheet_name]
        # the ws is a dictionary but the rows are tuples
        for row in ws.iter_rows(min_row=2):
            try:
                cathedra = Cathedra.query.filter_by(code=str(row[5].value)).all()[0]
                professor = Professor.query.filter_by(ci=str(row[6].value)).all()[0]
            except:
                return jsonify({"msg": "Catedra o profesor no existen"}), 400

            new_course = Course(
                title=row[0].value,
                code=row[1].value,
                init_date=row[2].value,
                finish_date=row[3].value,
                is_active=row[4].value,
                cathedra_id=cathedra.id,
                professor_id=professor.id
            )
            db.session.add(new_course)

    try:
        db.session.commit()
    except:
        db.session.rollback()
        return jsonify({"msg": "Hubo un problema creando el curso"}), 500

    return jsonify({"msg": "Se anadireron los cursos del archivo"}), 200

# endpoints for the evaluation
@app.route("/evaluation", methods=["POST"])
def create_evaluation():
    '''
        Creates an evaluation
    '''
    data = request.json

    new_evaluation = Evaluation(
        name=data["name"],
        percentage=data["percentage"],
        course_id=data["course_id"]
    )
    db.session.add(new_evaluation)

    try:
        db.session.commit()
    except:
        db.session.rollback()
        return jsonify({"msg": "Hubo un problema creando la evaluacion"}), 500

    return jsonify(new_evaluation.serialize()), 200

# endpoints for grades
@app.route("/upload-grades", methods=["POST"])
def upload_grades_file():
    '''
        Upload the data of grades from file
    '''
    # wb = workbook 
    try: 
        myFile = request.files["myFile"]
        wb = load_workbook(myFile)
    except: 
        return jsonify({"msg": "Hubo un problema abriendo el archivo"}), 500
    
    for sheet_name in wb.sheetnames:
        # ws = worksheet
        ws = wb[sheet_name]
        # the ws is a dictionary but the rows are tuples
        for row in ws.iter_rows(min_row=2):
            course_code = str(row[0].value)
            student_ci = str(row[1].value)
            evaluation_name = row[2].value
            evaluation_percentage = row[3].value
            grade_value = row[4].value
            # get the course by code
            course = Course.query.filter_by(code=course_code).all()[0]
            # get the student by ci
            student = Student.query.filter_by(ci=student_ci).all()[0]
            # get the inscription by the student id
            inscription = Inscription.query.filter_by(student_id=student.id).all()[0]
            # create the evaluation
            new_evaluation = Evaluation(name=evaluation_name, percentage=evaluation_percentage, course_id=course.id)
            db.session.add(new_evaluation)
            try: 
                db.session.commit()
            except:
                db.session.rollback()
                return jsonify({"msg": "Hubo un problema creando la evaluacion"}), 500
            # create the grade
            new_grade = Grade(value=grade_value, evaluation_id=new_evaluation.id, inscription_id=inscription.id)
            db.session.add(new_grade)
            try:
                db.session.commit()
            except:
                db.session.rollback()
                return jsonify({"msg": "Hubo un rpblema creando la nota"}), 500
    
    return jsonify({"msg": "Se anadireron los notas del archivo"}), 200

# this only runs if `$ python src/main.py` is executed
if __name__ == '__main__':
    PORT = int(os.environ.get('PORT', 4000))
    app.run(host='0.0.0.0', port=PORT, debug=False)
